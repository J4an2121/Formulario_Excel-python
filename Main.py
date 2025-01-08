import tkinter as tk #para crear la interfaz gráfica de usuario
from tkinter import messagebox #Sirve para mostrar cuadros de diálogo 
from openpyxl import Workbook, load_workbook   #Permite trabajar con archivos de Excel openpyxl
import re #Se usa para validar el formato de datos

# Intentar cargar el archivo existente o crear uno nuevo si no existe
try:
    wb = load_workbook('datos.xlsx')
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Edad", "Email", "Telefono", "Direccion"])  # Encabezados de las columnas

def guardar_datos():
    # Guardar información del formulario
    Nombre = entry_Nombre.get()
    Edad = entry_Edad.get()
    Email = entry_Email.get()
    Telefono = entry_Telefono.get()
    Direccion = entry_Direccion.get()

    # Validar que todos los campos estén llenos
    if not Nombre or not Edad or not Email or not Telefono or not Direccion:
        messagebox.showwarning("ADVERTENCIA", "Todos los campos son obligatorios")
        return

    # Validar que Edad y Teléfono sean números enteros
    try:
        Edad = int(Edad)
        Telefono = int(Telefono)
    except ValueError:
        messagebox.showwarning("ADVERTENCIA", "Edad y Teléfono deben ser números")
        return

    # Validar formato de email
    if not re.match(r"[^@]+@[^@]+\.[^@]+", Email):  # Validar patrón básico de email
        messagebox.showwarning(title="Advertencia", message="Email no válido")
        return

    # Guardar información en Excel
    ws.append([Nombre, Edad, Email, Telefono, Direccion])
    wb.save('datos.xlsx')
    messagebox.showinfo("Información", "Datos guardados correctamente")


    #Elimar informacion cuando se escribe 
    
    entry_Nombre.delete(0,tk.END)
    entry_Edad.delete(0,tk.END)
    entry_Email.delete(0,tk.END)
    entry_Telefono.delete(0,tk.END)
    entry_Direccion.delete(0,tk.END)


# Crear ventana principal
root = tk.Tk()
root.title("Formulario de Entrada de Datos")
root.configure(bg='#4b6587')  # Color de fondo
root.columnconfigure(1, weight=1)  # Permitir que las columnas crezcan dinámicamente

# Estilos
label_style = {"bg": '#4B6587', "fg": "white", "anchor": "w"}  # Estilo de etiquetas
entry_style = {"bg": '#D3D3D3', "fg": "black"}  # Estilo de campos de entrada

# Campos del formulario
label_Nombre = tk.Label(root, text="Nombre", **label_style)
label_Nombre.grid(row=0, column=0, padx=10, pady=5, sticky="w")
entry_Nombre = tk.Entry(root, **entry_style)
entry_Nombre.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

label_Edad = tk.Label(root, text="Edad", **label_style)
label_Edad.grid(row=1, column=0, padx=10, pady=5, sticky="w")
entry_Edad = tk.Entry(root, **entry_style)
entry_Edad.grid(row=1, column=1, padx=10, pady=5, sticky="ew")

label_Email = tk.Label(root, text="Email", **label_style)
label_Email.grid(row=2, column=0, padx=10, pady=5, sticky="w")
entry_Email = tk.Entry(root, **entry_style)
entry_Email.grid(row=2, column=1, padx=10, pady=5, sticky="ew")

label_Telefono = tk.Label(root, text="Teléfono", **label_style)
label_Telefono.grid(row=3, column=0, padx=10, pady=5, sticky="w")
entry_Telefono = tk.Entry(root, **entry_style)
entry_Telefono.grid(row=3, column=1, padx=10, pady=5, sticky="ew")

label_Direccion = tk.Label(root, text="Dirección", **label_style)
label_Direccion.grid(row=4, column=0, padx=10, pady=5, sticky="w")
entry_Direccion = tk.Entry(root, **entry_style)
entry_Direccion.grid(row=4, column=1, padx=10, pady=5, sticky="ew")

# Botón para guardar los datos
boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos, bg='#6D8299', fg='white')
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

# Ejecutar la ventana principal
root.mainloop()